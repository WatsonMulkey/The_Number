"""
FastAPI main application for The Number budgeting app.

This API provides REST endpoints for the Vue.js frontend,
wrapping the existing Python backend (database.py, calculator.py).
"""

import json
import os
import sys
import logging
from pathlib import Path
from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from typing import List, Optional
from dotenv import load_dotenv

# Configure logging
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Load environment variables FIRST - before any imports that depend on them
env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

# Add parent directory to path to import existing backend
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.database import EncryptedDatabase
from src.calculator import BudgetCalculator, WEEKLY_TO_MONTHLY
from src.import_expenses import import_expenses_from_file
from src.export_expenses import export_to_csv, export_to_excel
from api.models import (
    ExpenseCreate, ExpenseResponse, ExpenseUpdate, TransactionCreate, TransactionResponse,
    BudgetModeConfig, BudgetNumberResponse, ImportExpensesResponse, ErrorResponse,
    UserRegister, UserLogin, UserResponse, TokenResponse,
    ForgotPasswordRequest, ForgotPasswordResponse, ResetPasswordRequest, ResetPasswordResponse,
    PoolToggleRequest, PoolAddRequest, PoolSetRequest, PoolResponse
)
from api.auth import (
    hash_password, verify_password, create_access_token, get_current_user_id, get_admin_user_id,
    check_rate_limit, generate_reset_token, verify_reset_token, invalidate_reset_token
)

# Create FastAPI app
app = FastAPI(
    title="The Number API",
    description="REST API for The Number budgeting app",
    version="0.9.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc"
)

# Configure CORS
# Read allowed origins from environment variable
# For development: defaults to localhost on any port
# For production: set CORS_ORIGINS to your frontend URL(s)
cors_origins = os.getenv("CORS_ORIGINS", "http://localhost:5173,http://localhost:5174,http://localhost:5175,http://localhost:5176")
allowed_origins = [origin.strip() for origin in cors_origins.split(",")]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Startup validation for required environment variables
@app.on_event("startup")
async def validate_environment():
    """Validate that all required environment variables are set."""
    required_vars = ["DB_ENCRYPTION_KEY", "JWT_SECRET_KEY"]
    missing = [var for var in required_vars if not os.getenv(var)]

    if missing:
        error_msg = (
            f"Missing required environment variables: {', '.join(missing)}\n"
            f"Copy .env.example to .env and fill in all values.\n"
            f"See SETUP_GUIDE.md for instructions."
        )
        raise RuntimeError(error_msg)

    logger.info("Environment validation passed - encryption and auth configured")


# Dependency: Get database instance
def get_db() -> EncryptedDatabase:
    """
    Dependency that provides a database instance.

    Uses the encryption key from environment variables.

    Database path priority:
    1. DB_PATH environment variable
    2. /data/budget.db (production - Fly.io persistent volume)
    3. api/budget.db (development fallback)
    """
    encryption_key = os.getenv("DB_ENCRYPTION_KEY")
    if not encryption_key:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Database encryption key not configured"
        )

    # Determine database path
    db_path = os.getenv("DB_PATH")
    if not db_path:
        # Check if running in production (Fly.io has /data volume)
        if Path("/data").exists():
            db_path = "/data/budget.db"
        else:
            # Development: use local path
            db_path = str(Path(__file__).parent / "budget.db")

    return EncryptedDatabase(db_path=db_path, encryption_key=encryption_key)


# Root endpoint
@app.get("/")
async def root():
    """Root endpoint - API info."""
    return {
        "app": "The Number API",
        "version": "0.9.0",
        "docs": "/api/docs",
        "status": "running"
    }


# Health check
@app.get("/health")
async def health_check():
    """
    Health check endpoint.

    Verifies that critical environment variables are configured.
    """
    # Verify critical env vars exist
    has_db_key = bool(os.getenv("DB_ENCRYPTION_KEY"))
    has_jwt_key = bool(os.getenv("JWT_SECRET_KEY"))

    # Determine overall status
    is_healthy = has_db_key and has_jwt_key

    return {
        "status": "healthy" if is_healthy else "degraded",
        "encryption_configured": has_db_key,
        "auth_configured": has_jwt_key
    }


# ============================================================================
# BUDGET & "THE NUMBER" ENDPOINTS
# ============================================================================

@app.get("/api/number", response_model=BudgetNumberResponse)
async def get_the_number(
    response: Response,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Get "The Number" - your daily spending limit.

    This is the main feature of the app!
    Requires authentication.

    Note: This endpoint uses the user's configured timezone for date calculations.
    If no timezone is set, defaults to America/Denver (MST).
    """
    # Prevent caching - this data must always be fresh
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"

    try:
        # Get budget mode configuration for this user
        budget_mode = db.get_setting("budget_mode", user_id)

        if not budget_mode:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Budget mode not configured. Please configure your budget first."
            )

        # Get user's timezone for date calculations (default: MST)
        user_timezone = db.get_setting("user_timezone", user_id)

        # Load expenses into calculator
        calc = BudgetCalculator()
        expenses = db.get_expenses(user_id)
        for exp in expenses:
            calc.add_expense(exp["name"], exp["amount"], exp["is_fixed"],
                             exp.get("frequency", "monthly"))

        # Calculate "The Number" based on mode
        if budget_mode == "paycheck":
            from datetime import datetime, timedelta
            from api.utils.dates import get_user_today

            monthly_income = db.get_setting("monthly_income", user_id)

            # Try new approach: calculate from next_payday_date
            next_payday_str = db.get_setting("next_payday_date", user_id)
            pay_frequency = db.get_setting("pay_frequency_days", user_id) or 14

            if next_payday_str:
                next_payday = datetime.fromisoformat(next_payday_str).date()
                # Use user's timezone for "today" instead of server time
                today = get_user_today(user_timezone)

                # Auto-roll to next pay period if payday has passed
                while next_payday <= today:
                    # BEFORE advancing, calculate leftover from the previous cycle
                    # This enables the pool feature
                    previous_cycle_start = next_payday - timedelta(days=pay_frequency)

                    # Check if we already processed this payday (avoid double-counting)
                    last_processed = db.get_setting("last_processed_payday", user_id)
                    if last_processed != next_payday.isoformat():
                        # Calculate leftover for this cycle (normalize weekly→monthly)
                        total_expenses = sum(
                            exp["amount"] * WEEKLY_TO_MONTHLY if exp.get("frequency") == "weekly" else exp["amount"]
                            for exp in expenses
                        )
                        cycle_days = pay_frequency

                        # Get transactions during that cycle (UTC-aware boundaries)
                        from api.utils.dates import date_to_utc
                        from datetime import time as dt_time
                        transactions_total = db.get_transactions_sum_for_period(
                            user_id,
                            date_to_utc(previous_cycle_start, dt_time.min, user_timezone),
                            date_to_utc(next_payday, dt_time.min, user_timezone)
                        )

                        # Calculate what the budget was for that cycle (pro-rate monthly to cycle)
                        avg_days_per_month = 30.44
                        remaining_for_cycle = (monthly_income - total_expenses) * (cycle_days / avg_days_per_month)
                        expected_budget = remaining_for_cycle

                        # Leftover = what we had - what we spent
                        leftover = expected_budget - transactions_total

                        if leftover > 0:
                            # Accumulate pending contribution (handles multiple missed paydays)
                            current_pending = float(db.get_setting("pending_pool_contribution", user_id) or 0)
                            db.set_setting("pending_pool_contribution", current_pending + leftover, user_id)

                        # Track that we processed this payday
                        db.set_setting("last_processed_payday", next_payday.isoformat(), user_id)

                    # Advance to next payday
                    next_payday = next_payday + timedelta(days=pay_frequency)
                    # Update stored date for next time
                    db.set_setting("next_payday_date", next_payday.isoformat(), user_id)

                days_until_paycheck = (next_payday - today).days
            else:
                # Legacy: use static days_until_paycheck
                days_until_paycheck = db.get_setting("days_until_paycheck", user_id)

            result = calc.calculate_paycheck_mode(
                monthly_income=monthly_income,
                days_until_paycheck=days_until_paycheck,
                pay_frequency_days=int(pay_frequency)
            )

            # FOI-137: Adjust remaining_money for cycle-to-date spending.
            # The calculator returns the FULL cycle pro-rated budget. Without this,
            # the_number rises mid-cycle because full_budget / fewer_days = higher number,
            # ignoring what was already spent. Subtract prior days' spending so the_number
            # reflects the actual remaining money divided by actual remaining days.
            if next_payday_str and days_until_paycheck and days_until_paycheck > 0:
                from api.utils.dates import date_to_utc
                from datetime import time as dt_time

                cycle_start_date = next_payday - timedelta(days=int(pay_frequency))
                _today_spend = db.get_total_spending_today(user_id, user_timezone)
                # Use UTC-aware boundaries (matches get_total_spending_today's approach)
                # to avoid timezone mismatch near midnight for non-UTC users
                _total_cycle = db.get_transactions_sum_for_period(
                    user_id,
                    date_to_utc(cycle_start_date, dt_time.min, user_timezone),
                    date_to_utc(today + timedelta(days=1), dt_time.min, user_timezone)
                )
                _prior_days_spending = max(0, _total_cycle - _today_spend)
                result["remaining_money"] -= _prior_days_spending
                result["daily_limit"] = max(0, result["remaining_money"] / days_until_paycheck)

        else:  # fixed_pool
            total_money = db.get_setting("total_money", user_id)
            target_end_date_str = db.get_setting("target_end_date", user_id)
            daily_spending_limit = db.get_setting("daily_spending_limit", user_id)

            # Parse target_end_date if it exists
            target_end_date = None
            if target_end_date_str:
                from datetime import datetime
                target_end_date = datetime.fromisoformat(target_end_date_str)
            result = calc.calculate_fixed_pool_mode(
                total_money=total_money,
                target_end_date=target_end_date,
                daily_spending_limit=daily_spending_limit
            )


        # One-time pool reset: pool balances computed before 2026-03-16 used a buggy
        # formula that treated monthly surplus as per-cycle surplus, inflating the pool.
        # Reset pool and pending contributions so users start clean with corrected math.
        pool_formula_fixed = db.get_setting("pool_formula_fixed", user_id)
        if not pool_formula_fixed:
            db.set_setting("pool_balance", 0, user_id)
            db.set_setting("pending_pool_contribution", 0, user_id)
            db.set_setting("pool_formula_fixed", True, user_id)

        # Get pool settings
        pool_enabled = db.get_setting("pool_enabled", user_id) == True
        pool_balance = float(db.get_setting("pool_balance", user_id) or 0)
        pending_pool = float(db.get_setting("pending_pool_contribution", user_id) or 0)

        # Base daily limit from calculator
        base_daily_limit = result["daily_limit"]
        days_remaining = result.get("days_remaining")
        remaining_money = result.get("remaining_money", 0)

        # Defensive fallback: if remaining_money is absent from the result dict,
        # reconstruct it to prevent pool math corruption (FOI-103)
        if "remaining_money" not in result and base_daily_limit > 0 and days_remaining and days_remaining > 0:
            remaining_money = base_daily_limit * days_remaining

        # If pool is enabled and has balance, factor it into the daily budget
        the_number = base_daily_limit
        if pool_enabled and pool_balance > 0 and days_remaining and days_remaining > 0:
            # Add pool to remaining money, recalculate daily limit
            remaining_with_pool = remaining_money + pool_balance
            the_number = remaining_with_pool / days_remaining

        # Get today's spending (using user's timezone for day boundaries)
        today_spending = db.get_total_spending_today(user_id, user_timezone)
        remaining_today = the_number - today_spending
        is_over_budget = remaining_today < 0

        # Calculate adjusted/tomorrow daily budget
        adjusted_daily_budget = None
        original_daily_budget = None
        tomorrow_daily_budget = None

        # Only calculate if we have more than 1 day remaining
        if days_remaining and days_remaining > 1:
            # Calculate what tomorrow's budget would be based on today's spending
            effective_remaining = remaining_money + pool_balance if pool_enabled else remaining_money
            remaining_after_today = effective_remaining - today_spending
            tomorrow_budget = remaining_after_today / (days_remaining - 1)

            if is_over_budget:
                # Overspend: show adjusted budget (what you now have to work with)
                original_daily_budget = the_number
                if tomorrow_budget > 0:
                    adjusted_daily_budget = round(tomorrow_budget, 2)
            else:
                # Underspend or on-budget: show tomorrow's preview
                # Only show if tomorrow's budget differs meaningfully from today's
                if tomorrow_budget > 0:
                    tomorrow_daily_budget = round(tomorrow_budget, 2)

        return BudgetNumberResponse(
            the_number=the_number,
            mode=budget_mode,
            total_income=result.get("total_income"),
            total_money=result.get("total_money"),
            total_expenses=result.get("total_expenses", result.get("monthly_expenses", 0)),
            remaining_money=result.get("remaining_money"),
            days_remaining=days_remaining,
            today_spending=today_spending,
            remaining_today=remaining_today,
            is_over_budget=is_over_budget,
            adjusted_daily_budget=adjusted_daily_budget,
            original_daily_budget=original_daily_budget,
            tomorrow_daily_budget=tomorrow_daily_budget,
            pool_balance=pool_balance,
            pool_enabled=pool_enabled,
            pending_pool_contribution=pending_pool if pending_pool > 0 else None
        )

    except HTTPException:
        # Re-raise HTTP exceptions as-is (don't wrap in 500)
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error calculating budget: {str(e)}"
        )


@app.post("/api/budget/configure", status_code=status.HTTP_200_OK)
async def configure_budget(
    config: BudgetModeConfig,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Configure budget mode and settings for the authenticated user.
    Requires authentication.
    """
    logger.info(f"Budget configured for user {user_id} in {config.mode} mode")
    try:
        # Validate configuration based on mode
        if config.mode == "paycheck":
            if not config.monthly_income:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Paycheck mode requires monthly_income"
                )

            # Require either next_payday_date (preferred) or days_until_paycheck (legacy)
            if not config.next_payday_date and not config.days_until_paycheck:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Paycheck mode requires next_payday_date or days_until_paycheck"
                )

            db.set_setting("monthly_income", config.monthly_income, user_id)

            # If next_payday_date provided, use it (new approach)
            if config.next_payday_date:
                db.set_setting("next_payday_date", config.next_payday_date.isoformat(), user_id)
                db.set_setting("pay_frequency_days", config.pay_frequency_days or 14, user_id)
                # Clear legacy setting
                db.set_setting("days_until_paycheck", None, user_id)
            else:
                # Legacy: use days_until_paycheck directly
                db.set_setting("days_until_paycheck", config.days_until_paycheck, user_id)

        elif config.mode == "fixed_pool":
            if config.total_money is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Fixed pool mode requires total_money"
                )
            db.set_setting("total_money", config.total_money, user_id)

            # Save fixed pool options (Option B and C)
            if config.target_end_date:
                db.set_setting("target_end_date", config.target_end_date.isoformat(), user_id)
            else:
                # Clear it if not provided
                db.set_setting("target_end_date", None, user_id)

            if config.daily_spending_limit:
                db.set_setting("daily_spending_limit", config.daily_spending_limit, user_id)
            else:
                # Clear it if not provided
                db.set_setting("daily_spending_limit", None, user_id)

        # Save budget mode
        db.set_setting("budget_mode", config.mode, user_id)

        # Save user timezone if provided (for correct day boundary calculations)
        if config.user_timezone:
            from api.utils.dates import validate_timezone
            validated_tz = validate_timezone(config.user_timezone)
            db.set_setting("user_timezone", validated_tz, user_id)

        return {"message": f"Budget configured successfully in {config.mode} mode"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error configuring budget: {str(e)}"
        )


@app.get("/api/budget/config")
async def get_budget_config(
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Get current budget configuration for the authenticated user.
    Requires authentication.
    """
    mode = db.get_setting("budget_mode", user_id)
    if not mode:
        return {"configured": False}

    config = {"configured": True, "mode": mode}

    if mode == "paycheck":
        config["monthly_income"] = db.get_setting("monthly_income", user_id)
        config["next_payday_date"] = db.get_setting("next_payday_date", user_id)
        config["pay_frequency_days"] = db.get_setting("pay_frequency_days", user_id)
        # Also return legacy field for backwards compatibility
        config["days_until_paycheck"] = db.get_setting("days_until_paycheck", user_id)
    else:
        config["total_money"] = db.get_setting("total_money", user_id)
        target_end_date = db.get_setting("target_end_date", user_id)
        daily_spending_limit = db.get_setting("daily_spending_limit", user_id)
        if target_end_date:
            config["target_end_date"] = target_end_date
        if daily_spending_limit:
            config["daily_spending_limit"] = daily_spending_limit

    return config


# ============================================================================
# EXPENSE ENDPOINTS
# ============================================================================

@app.get("/api/expenses", response_model=List[ExpenseResponse])
async def get_expenses(
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Get all expenses for the authenticated user.
    Requires authentication.
    """
    return db.get_expenses(user_id)


@app.post("/api/expenses", response_model=ExpenseResponse, status_code=status.HTTP_201_CREATED)
async def create_expense(
    expense: ExpenseCreate,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Create a new expense for the authenticated user.
    Requires authentication.
    """
    try:
        expense_id = db.add_expense(
            name=expense.name,
            amount=expense.amount,
            user_id=user_id,
            is_fixed=expense.is_fixed,
            frequency=expense.frequency
        )

        # Retrieve the created expense
        created_expense = db.get_expense_by_id(expense_id, user_id)
        return created_expense

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating expense: {str(e)}"
        )


@app.delete("/api/expenses/{expense_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_expense(
    expense_id: int,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Delete an expense for the authenticated user.
    Requires authentication.
    """
    try:
        db.delete_expense(expense_id, user_id)
        return None
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting expense: {str(e)}"
        )


@app.put("/api/expenses/{expense_id}", response_model=ExpenseResponse)
async def update_expense(
    expense_id: int,
    expense: ExpenseUpdate,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Update an expense for the authenticated user.
    Supports partial updates - only provided fields are updated.
    Requires authentication.
    """
    try:
        # Verify expense exists and belongs to user
        existing = db.get_expense_by_id(expense_id, user_id)
        if not existing:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Expense not found"
            )

        # Update expense (database method handles partial updates)
        db.update_expense(
            expense_id=expense_id,
            user_id=user_id,
            name=expense.name,
            amount=expense.amount,
            is_fixed=expense.is_fixed,
            frequency=expense.frequency
        )

        # Return updated expense
        updated = db.get_expense_by_id(expense_id, user_id)
        return updated

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error updating expense: {str(e)}"
        )


# SECURITY: File upload validation constants
MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10MB max file size
ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}
ALLOWED_CONTENT_TYPES = {
    'text/csv',
    'application/vnd.ms-excel',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/csv',
    'text/plain',  # Some systems send CSV as text/plain
}


@app.post("/api/expenses/import", response_model=ImportExpensesResponse)
async def import_expenses(
    file: UploadFile = File(...),
    replace: bool = False,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Import expenses from CSV or Excel file for the authenticated user.
    Requires authentication.

    - **file**: CSV or Excel file containing expenses (max 10MB, .csv or .xlsx only)
    - **replace**: If True, replaces all existing expenses. If False, adds to existing.
    """
    try:
        # SECURITY FIX: Validate file before processing
        # 1. Validate file extension
        if file.filename:
            file_ext = Path(file.filename).suffix.lower()
            if file_ext not in ALLOWED_EXTENSIONS:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid file type. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
                )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Filename is required"
            )

        # 2. Validate content type
        if file.content_type and file.content_type not in ALLOWED_CONTENT_TYPES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid content type: {file.content_type}. Expected CSV or Excel file."
            )

        # 3. Validate file size by reading content
        # Read file content to check size (also needed for processing)
        content = await file.read()
        if len(content) > MAX_FILE_SIZE_BYTES:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail=f"File too large. Maximum size is {MAX_FILE_SIZE_BYTES // (1024 * 1024)}MB"
            )

        if len(content) == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File is empty"
            )

        # Reset file position for processing (create a BytesIO object)
        from io import BytesIO
        file_obj = BytesIO(content)

        # Import expenses using existing backend
        expenses, errors = import_expenses_from_file(file_obj)

        # Replace existing expenses if requested
        if replace:
            existing = db.get_expenses(user_id)
            for exp in existing:
                db.delete_expense(exp["id"], user_id)

        # Add imported expenses
        imported_count = 0
        for exp in expenses:
            try:
                db.add_expense(exp["name"], exp["amount"], user_id, exp["is_fixed"])
                imported_count += 1
            except Exception as e:
                errors.append(f"Failed to import {exp['name']}: {str(e)}")

        return ImportExpensesResponse(
            imported_count=imported_count,
            errors=errors
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error importing expenses: {str(e)}"
        )


@app.get("/api/expenses/export/{format}")
async def export_expenses(
    format: str,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Export expenses to CSV or Excel for the authenticated user.
    Requires authentication.

    - **format**: Either 'csv' or 'excel'
    """
    try:
        expenses = db.get_expenses(user_id)

        if format.lower() == "csv":
            file_path = export_to_csv(expenses)
        elif format.lower() in ["excel", "xlsx"]:
            file_path = export_to_excel(expenses)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Format must be 'csv' or 'excel'"
            )

        return FileResponse(
            file_path,
            media_type="application/octet-stream",
            filename=Path(file_path).name
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exporting expenses: {str(e)}"
        )


# ============================================================================
# TRANSACTION ENDPOINTS
# ============================================================================

@app.get("/api/transactions", response_model=List[TransactionResponse])
async def get_transactions(
    limit: Optional[int] = 20,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Get recent transactions for the authenticated user.
    Requires authentication.
    """
    return db.get_transactions(user_id, limit=limit)


@app.post("/api/transactions", response_model=TransactionResponse, status_code=status.HTTP_201_CREATED)
async def create_transaction(
    transaction: TransactionCreate,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Record a spending transaction for the authenticated user.
    Requires authentication.
    """
    try:
        txn_id = db.add_transaction(
            amount=transaction.amount,
            description=transaction.description,
            user_id=user_id,
            date=transaction.date,
            category=transaction.category
        )

        # Retrieve all transactions and find the one we just created
        transactions = db.get_transactions(user_id, limit=1)
        return transactions[0] if transactions else None

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creating transaction: {str(e)}"
        )


@app.delete("/api/transactions/{transaction_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_transaction(
    transaction_id: int,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Delete a transaction for the authenticated user.
    Requires authentication.
    """
    try:
        db.delete_transaction(transaction_id, user_id)
        return None
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting transaction: {str(e)}"
        )


# ============================================================================
# POOL ENDPOINTS
# ============================================================================

@app.post("/api/pool/accept", response_model=PoolResponse)
async def accept_pool_contribution(
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Accept pending pool contribution from payday rollover.

    Moves the pending contribution to the pool balance.
    Requires authentication.
    """
    pending = float(db.get_setting("pending_pool_contribution", user_id) or 0)
    if pending <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No pending contribution to accept"
        )

    current_balance = float(db.get_setting("pool_balance", user_id) or 0)
    new_balance = current_balance + pending

    db.set_setting("pool_balance", new_balance, user_id)
    db.set_setting("pending_pool_contribution", 0, user_id)

    logger.info(f"User {user_id} accepted pool contribution: ${pending:.2f}, new balance: ${new_balance:.2f}")

    return PoolResponse(pool_balance=new_balance)


@app.post("/api/pool/decline")
async def decline_pool_contribution(
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Decline pending pool contribution.

    Clears the pending contribution without adding to pool.
    Requires authentication.
    """
    pending = float(db.get_setting("pending_pool_contribution", user_id) or 0)
    db.set_setting("pending_pool_contribution", 0, user_id)

    logger.info(f"User {user_id} declined pool contribution: ${pending:.2f}")

    return {"status": "declined", "amount_declined": pending}


@app.post("/api/pool/toggle", response_model=PoolResponse)
async def toggle_pool(
    request: PoolToggleRequest,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Toggle pool enabled/disabled for daily budget calculation.

    When enabled, pool balance is factored into the daily budget.
    Requires authentication.
    """
    db.set_setting("pool_enabled", request.enabled, user_id)
    pool_balance = float(db.get_setting("pool_balance", user_id) or 0)

    logger.info(f"User {user_id} toggled pool: enabled={request.enabled}")

    return PoolResponse(pool_balance=pool_balance)


@app.post("/api/pool/add", response_model=PoolResponse)
async def add_to_pool(
    request: PoolAddRequest,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Manually add money to the pool.

    Requires authentication.
    """
    current_balance = float(db.get_setting("pool_balance", user_id) or 0)
    new_balance = current_balance + request.amount

    db.set_setting("pool_balance", new_balance, user_id)

    logger.info(f"User {user_id} added ${request.amount:.2f} to pool, new balance: ${new_balance:.2f}")

    return PoolResponse(pool_balance=new_balance)


@app.post("/api/pool/set", response_model=PoolResponse)
async def set_pool_balance(
    request: PoolSetRequest,
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Set pool balance to an exact value.

    Allows the user to adjust their pool balance directly (e.g., after dipping into savings).
    Requires authentication.
    """
    db.set_setting("pool_balance", request.balance, user_id)

    logger.info(f"User {user_id} set pool balance to ${request.balance:.2f}")

    return PoolResponse(pool_balance=request.balance)


@app.get("/api/pool")
async def get_pool_status(
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Get current pool status.

    Returns pool balance, enabled status, and any pending contribution.
    Requires authentication.
    """
    pool_balance = float(db.get_setting("pool_balance", user_id) or 0)
    pool_enabled = db.get_setting("pool_enabled", user_id) == True
    pending = float(db.get_setting("pending_pool_contribution", user_id) or 0)

    return {
        "pool_balance": pool_balance,
        "pool_enabled": pool_enabled,
        "pending_pool_contribution": pending if pending > 0 else None
    }


# ============================================================================
# AUTHENTICATION ENDPOINTS
# ============================================================================

@app.post("/api/auth/register", response_model=TokenResponse, status_code=status.HTTP_201_CREATED)
async def register(
    user_data: UserRegister,
    request: Request,
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Register a new user.

    Creates a new user account and returns an access token.
    Rate limited: 5 requests per 60 seconds per IP.
    """
    # Rate limiting: 5 requests per minute
    check_rate_limit(request, max_requests=5, window_seconds=60)

    # --- BETA GATING (remove after beta) ---
    beta_codes = os.environ.get("BETA_INVITE_CODES", "")
    if beta_codes:
        valid_codes = [c.strip().upper() for c in beta_codes.split(",") if c.strip()]
        provided = (user_data.invite_code or "").strip().upper()
        if not provided or provided not in valid_codes:
            raise HTTPException(status_code=400, detail="Invalid invite code")
    # --- END BETA GATING ---

    try:
        # Check if username already exists
        existing_user = db.get_user_by_username(user_data.username)
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Username already exists"
            )

        # Hash password
        password_hash = hash_password(user_data.password)

        # Create user
        user_id = db.create_user(
            username=user_data.username,
            password_hash=password_hash,
            email=user_data.email
        )

        # Store user timezone if provided (auto-detected by frontend)
        if user_data.timezone:
            from api.utils.dates import validate_timezone
            validated_tz = validate_timezone(user_data.timezone)
            db.set_setting("user_timezone", validated_tz, user_id)
            db.set_setting("timezone_source", "auto", user_id)

        # Get created user
        user = db.get_user_by_id(user_id)

        # Create access token
        access_token = create_access_token(data={"user_id": user_id})

        admin_env = os.getenv("ADMIN_USER_ID")
        return TokenResponse(
            access_token=access_token,
            token_type="bearer",
            user=UserResponse(
                id=user["id"],
                username=user["username"],
                email=user["email"],
                created_at=user["created_at"],
                is_admin=bool(admin_env) and str(user["id"]) == admin_env
            )
        )

    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error registering user: {str(e)}"
        )


@app.post("/api/auth/login", response_model=TokenResponse)
async def login(
    credentials: UserLogin,
    request: Request,
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Login with username and password.

    Returns an access token on successful authentication.
    Rate limited: 5 requests per 60 seconds per IP.
    """
    # Rate limiting: 5 requests per minute to prevent brute force
    check_rate_limit(request, max_requests=5, window_seconds=60)

    try:
        # Get user by username
        user = db.get_user_by_username(credentials.username)
        if not user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid username or password"
            )

        # Verify password
        if not verify_password(credentials.password, user["password_hash"]):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid username or password"
            )

        # Update user timezone on login only if source is auto-detected (not manually set)
        if credentials.timezone:
            tz_source = db.get_setting("timezone_source", user["id"])
            if tz_source != "manual":
                from api.utils.dates import validate_timezone
                validated_tz = validate_timezone(credentials.timezone)
                db.set_setting("user_timezone", validated_tz, user["id"])
                db.set_setting("timezone_source", "auto", user["id"])

        # Create access token
        access_token = create_access_token(data={"user_id": user["id"]})

        admin_env = os.getenv("ADMIN_USER_ID")
        return TokenResponse(
            access_token=access_token,
            token_type="bearer",
            user=UserResponse(
                id=user["id"],
                username=user["username"],
                email=user["email"],
                created_at=user["created_at"],
                is_admin=bool(admin_env) and str(user["id"]) == admin_env
            )
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error logging in: {str(e)}"
        )


@app.get("/api/auth/me", response_model=UserResponse)
async def get_current_user(
    user_id: int = Depends(get_current_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Get current authenticated user.

    Requires valid JWT token in Authorization header.
    """
    user = db.get_user_by_id(user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )

    # Record activity for DAU/WAU/MAU metrics (never fail auth for metrics)
    try:
        from datetime import date
        db.record_user_activity(user_id, date.today().isoformat())
    except Exception:
        pass

    admin_env = os.getenv("ADMIN_USER_ID")
    return UserResponse(
        id=user["id"],
        username=user["username"],
        email=user["email"],
        created_at=user["created_at"],
        is_admin=bool(admin_env) and str(user["id"]) == admin_env
    )


@app.post("/api/auth/logout")
async def logout():
    """
    Logout endpoint.

    Since we're using JWT, logout is handled client-side by removing the token.
    This endpoint exists for consistency and future server-side logout logic.
    """
    return {"message": "Logged out successfully"}


@app.post("/api/auth/forgot-password", response_model=ForgotPasswordResponse)
async def forgot_password(
    request: Request,
    forgot_request: ForgotPasswordRequest
):
    """
    Request a password reset token.

    Generates a password reset token for the given username.
    The token is valid for 1 hour.

    Rate limited to prevent abuse.
    """
    # Apply rate limiting
    check_rate_limit(request, max_requests=5, window_seconds=300)  # 5 requests per 5 minutes

    # Verify username exists
    db = EncryptedDatabase()
    user = db.get_user_by_username(forgot_request.username)

    if not user:
        # Don't reveal if username exists or not (security best practice)
        # Return success even if user doesn't exist
        pass

    # Generate reset token
    reset_token = generate_reset_token(forgot_request.username)

    return ForgotPasswordResponse(
        reset_token=reset_token,
        message="Password reset token generated. Use this token to reset your password within 1 hour.",
        expires_in=3600  # 1 hour in seconds
    )


@app.post("/api/auth/reset-password", response_model=ResetPasswordResponse)
async def reset_password(
    request: Request,
    reset_request: ResetPasswordRequest
):
    """
    Reset password using a reset token.

    Validates the reset token and updates the user's password.
    The token is invalidated after successful password reset.
    """
    # Verify reset token
    username = verify_reset_token(reset_request.reset_token)

    if not username:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid or expired reset token"
        )

    # Update password
    db = EncryptedDatabase()
    user = db.get_user_by_username(username)

    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )

    # Hash new password
    hashed_password = hash_password(reset_request.new_password)

    # Update password in database
    db.update_user_password(user["id"], hashed_password)

    # Invalidate the reset token
    invalidate_reset_token(reset_request.reset_token)

    return ResetPasswordResponse(
        message="Password has been reset successfully. You can now login with your new password."
    )


# ============================================================================
# ADMIN / BACKUP ENDPOINTS
# ============================================================================

@app.post("/api/admin/backup")
async def create_backup(user_id: int = Depends(get_current_user_id)):
    """
    Create a manual database backup.

    Creates a timestamped backup in the backups/manual directory.
    Requires authentication.
    """
    import sqlite3
    import shutil
    from datetime import datetime
    from pathlib import Path as PathLib

    try:
        # Use absolute paths from the file location
        api_dir = PathLib(__file__).parent
        project_root = api_dir.parent

        db_path = api_dir / "budget.db"
        backup_dir = project_root / "backups" / "manual"

        # Create backup directory if needed
        backup_dir.mkdir(parents=True, exist_ok=True)

        # Generate timestamped backup filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"budget_backup_{timestamp}.db"
        backup_path = backup_dir / backup_name

        # Create backup using SQLite's backup API
        src = sqlite3.connect(str(db_path))
        dst = sqlite3.connect(str(backup_path))

        with dst:
            src.backup(dst)

        src.close()
        dst.close()

        return {
            "success": True,
            "backup_path": str(backup_path),
            "backup_filename": backup_name,
            "created_at": datetime.now().isoformat()
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Backup failed: {str(e)}"
        )


@app.get("/api/admin/backups")
async def list_backups(user_id: int = Depends(get_current_user_id)):
    """
    List all available database backups.

    Returns backups from both manual and automatic directories,
    sorted by creation time (newest first).
    """
    from pathlib import Path
    from datetime import datetime

    # Use absolute path to backups directory
    backup_root = Path(__file__).parent.parent / "backups"
    if not backup_root.exists():
        return {"backups": []}

    all_backups = []

    # Scan both manual and automatic directories
    for subdir in ["manual", "automatic"]:
        subdir_path = backup_root / subdir
        if subdir_path.exists():
            for backup in subdir_path.glob("budget_backup_*.db"):
                stat = backup.stat()
                all_backups.append({
                    "filename": backup.name,
                    "path": str(backup),
                    "type": subdir,
                    "size": stat.st_size,
                    "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
                })

    # Sort by creation time (newest first)
    all_backups.sort(key=lambda b: b["created_at"], reverse=True)

    # Limit to most recent 20
    return {"backups": all_backups[:20]}


@app.get("/api/export/{format}")
async def export_budget_data(format: str, user_id: int = Depends(get_current_user_id)):
    """
    Export all budget data (config, expenses, transactions) in CSV or Excel format.

    Supported formats: 'csv', 'excel'
    """
    import csv
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse

    if format not in ['csv', 'excel']:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'excel'")

    db = get_db()

    # Get budget configuration via get_setting (settings are encrypted)
    setting_keys = [
        "budget_mode", "monthly_income", "next_payday_date", "pay_frequency_days",
        "days_until_paycheck", "total_money", "target_end_date", "daily_spending_limit",
        "user_timezone", "pool_enabled", "pool_balance", "pending_pool_contribution",
    ]
    settings = {}
    for key in setting_keys:
        val = db.get_setting(key, user_id)
        if val is not None:
            settings[key] = val

    # Get expenses via get_expenses (returns list of dicts)
    expenses_raw = db.get_expenses(user_id)

    # Get transactions via get_transactions (returns list of dicts)
    transactions_raw = db.get_transactions(user_id)

    if format == 'csv':
        # Create CSV with multiple sections
        output = io.StringIO()
        writer = csv.writer(output)

        # Budget Configuration Section
        writer.writerow(['BUDGET CONFIGURATION'])
        writer.writerow(['Setting', 'Value'])
        for key, value in settings.items():
            writer.writerow([key, value])
        writer.writerow([])  # Blank line

        # Expenses Section
        writer.writerow(['MONTHLY EXPENSES'])
        writer.writerow(['Name', 'Amount', 'Type', 'Frequency'])
        for exp in expenses_raw:
            exp_type = 'Fixed' if exp.get('is_fixed') else 'Variable'
            writer.writerow([exp['name'], f'${exp["amount"]:.2f}', exp_type, exp.get('frequency', 'monthly')])
        writer.writerow([])

        # Transactions Section
        writer.writerow(['TRANSACTIONS'])
        writer.writerow(['Date', 'Amount', 'Description', 'Category'])
        for txn in transactions_raw:
            category = txn.get('category', '') or ''
            writer.writerow([txn.get('date', ''), f'${txn["amount"]:.2f}', txn.get('description', ''), category])

        # Return CSV file
        output.seek(0)
        filename = f"budget_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    else:  # Excel format
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="Excel export requires openpyxl. Please install it."
            )

        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()

        # Sheet 1: Budget Configuration
        ws_config = wb.active
        ws_config.title = "Budget Config"
        ws_config['A1'] = 'Budget Configuration'
        ws_config['A1'].font = Font(bold=True, size=14)
        ws_config['A3'] = 'Setting'
        ws_config['B3'] = 'Value'
        ws_config['A3'].font = Font(bold=True)
        ws_config['B3'].font = Font(bold=True)

        row = 4
        for key, value in settings.items():
            ws_config[f'A{row}'] = key
            ws_config[f'B{row}'] = value
            row += 1

        # Sheet 2: Expenses
        ws_expenses = wb.create_sheet("Expenses")
        ws_expenses['A1'] = 'Monthly Expenses'
        ws_expenses['A1'].font = Font(bold=True, size=14)
        ws_expenses['A3'] = 'Name'
        ws_expenses['B3'] = 'Amount'
        ws_expenses['C3'] = 'Type'
        ws_expenses['D3'] = 'Frequency'
        for col in ['A3', 'B3', 'C3', 'D3']:
            ws_expenses[col].font = Font(bold=True)

        row = 4
        total = 0
        for exp in expenses_raw:
            exp_type = 'Fixed' if exp.get('is_fixed') else 'Variable'
            ws_expenses[f'A{row}'] = exp['name']
            ws_expenses[f'B{row}'] = exp['amount']
            ws_expenses[f'B{row}'].number_format = '$#,##0.00'
            ws_expenses[f'C{row}'] = exp_type
            ws_expenses[f'D{row}'] = exp.get('frequency', 'monthly')
            total += exp['amount']
            row += 1

        # Add total
        ws_expenses[f'A{row}'] = 'TOTAL'
        ws_expenses[f'A{row}'].font = Font(bold=True)
        ws_expenses[f'B{row}'] = total
        ws_expenses[f'B{row}'].number_format = '$#,##0.00'
        ws_expenses[f'B{row}'].font = Font(bold=True)

        # Sheet 3: Transactions
        ws_txns = wb.create_sheet("Transactions")
        ws_txns['A1'] = 'Transaction History'
        ws_txns['A1'].font = Font(bold=True, size=14)
        ws_txns['A3'] = 'Date'
        ws_txns['B3'] = 'Amount'
        ws_txns['C3'] = 'Description'
        ws_txns['D3'] = 'Category'
        for col in ['A3', 'B3', 'C3', 'D3']:
            ws_txns[col].font = Font(bold=True)

        row = 4
        for txn in transactions_raw:
            ws_txns[f'A{row}'] = txn.get('date', '')
            ws_txns[f'B{row}'] = txn['amount']
            ws_txns[f'B{row}'].number_format = '$#,##0.00'
            ws_txns[f'C{row}'] = txn.get('description', '')
            ws_txns[f'D{row}'] = txn.get('category', '') or ''
            row += 1

        # Auto-size columns
        for ws in [ws_config, ws_expenses, ws_txns]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"budget_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@app.get("/api/admin/backups/download/{filename}")
async def download_backup(filename: str, user_id: int = Depends(get_current_user_id)):
    """
    Download a specific backup file.

    Security: Validates filename to prevent directory traversal attacks.
    Returns the backup file as a downloadable attachment.
    """
    from pathlib import Path
    from fastapi.responses import FileResponse
    import os

    # Validate filename to prevent directory traversal
    if "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(
            status_code=400,
            detail="Invalid filename"
        )

    # Only allow .db files with correct naming pattern
    if not filename.startswith("budget_backup_") or not filename.endswith(".db"):
        raise HTTPException(
            status_code=400,
            detail="Invalid backup filename format"
        )

    # Use same path logic as create_backup
    api_dir = Path(__file__).parent
    project_root = api_dir.parent

    # Search for the backup in both manual and automatic directories
    backup_path = None
    for subdir in ["manual", "automatic"]:
        potential_path = project_root / "backups" / subdir / filename
        logger.info(f"Checking backup path: {potential_path}")
        logger.info(f"File exists: {potential_path.exists()}")
        if potential_path.exists():
            backup_path = potential_path
            logger.info(f"Found backup at: {backup_path}")
            break

    if not backup_path:
        logger.error(f"Backup file not found in {project_root / 'backups'}")
        raise HTTPException(
            status_code=404,
            detail="Backup file not found"
        )

    # Return file as download
    return FileResponse(
        path=str(backup_path),
        filename=filename,
        media_type="application/x-sqlite3",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


# ============================================================================
# ADMIN METRICS ENDPOINTS
# ============================================================================

@app.get("/api/admin/metrics")
async def get_admin_metrics(
    admin_id: int = Depends(get_admin_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    Aggregate usage metrics for admin dashboard. No PII in response.
    Protected by admin auth.
    """
    import sqlite3 as _sqlite3
    from datetime import date, timedelta

    today = date.today()
    week_ago = (today - timedelta(days=7)).isoformat()
    month_ago = (today - timedelta(days=30)).isoformat()

    with _sqlite3.connect(db.db_path) as conn:
        cursor = conn.cursor()

        # --- Growth ---
        cursor.execute("SELECT COUNT(*) FROM users")
        total_users = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM users WHERE created_at >= ?", (week_ago,))
        signups_this_week = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM users WHERE created_at >= ?", (month_ago,))
        signups_this_month = cursor.fetchone()[0]

        # --- Engagement (from user_activity table) ---
        cursor.execute("SELECT COUNT(DISTINCT user_id) FROM user_activity WHERE activity_date = ?", (today.isoformat(),))
        dau = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(DISTINCT user_id) FROM user_activity WHERE activity_date >= ?", (week_ago,))
        wau = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(DISTINCT user_id) FROM user_activity WHERE activity_date >= ?", (month_ago,))
        mau = cursor.fetchone()[0]

        cursor.execute("SELECT AVG(login_count) FROM user_activity WHERE activity_date = ?", (today.isoformat(),))
        avg_sessions = cursor.fetchone()[0] or 0

        # --- Feature Adoption (decrypt settings) ---
        cursor.execute("SELECT value FROM settings WHERE key = 'budget_mode'")
        mode_rows = cursor.fetchall()
        total_configured = len(mode_rows)
        paycheck_count = 0
        fixed_pool_count = 0
        for (encrypted_val,) in mode_rows:
            try:
                val = json.loads(db._decrypt(encrypted_val))
                if val == "paycheck":
                    paycheck_count += 1
                elif val == "fixed_pool":
                    fixed_pool_count += 1
            except Exception:
                pass

        cursor.execute("SELECT value FROM settings WHERE key = 'pool_enabled'")
        pool_rows = cursor.fetchall()
        pool_enabled_count = 0
        for (encrypted_val,) in pool_rows:
            try:
                val = json.loads(db._decrypt(encrypted_val))
                if val is True:
                    pool_enabled_count += 1
            except Exception:
                pass

        # --- Volume ---
        cursor.execute("SELECT COUNT(*) FROM transactions")
        total_transactions = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM expenses")
        total_expenses = cursor.fetchone()[0]

    return {
        "growth": {
            "total_users": total_users,
            "signups_this_week": signups_this_week,
            "signups_this_month": signups_this_month,
        },
        "engagement": {
            "dau": dau,
            "wau": wau,
            "mau": mau,
            "avg_sessions_per_day": round(avg_sessions, 1),
        },
        "depth": {
            "budget_configured_count": total_configured,
            "budget_configured_pct": round(total_configured / total_users * 100, 1) if total_users else 0,
            "paycheck_mode_count": paycheck_count,
            "fixed_pool_mode_count": fixed_pool_count,
            "pool_enabled_count": pool_enabled_count,
        },
        "volume": {
            "total_transactions": total_transactions,
            "total_expenses": total_expenses,
        },
    }


@app.get("/api/admin/trends")
async def get_admin_trends(
    admin_id: int = Depends(get_admin_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """30-day daily trends + week-over-week comparisons for admin dashboard."""
    import sqlite3 as _sqlite3
    from datetime import date, timedelta

    today = date.today()
    start = today - timedelta(days=29)

    # Build date range for gap-filling
    date_range = [(start + timedelta(days=i)).isoformat() for i in range(30)]

    with _sqlite3.connect(db.db_path) as conn:
        cursor = conn.cursor()

        # Daily active users from user_activity
        cursor.execute(
            "SELECT activity_date, COUNT(DISTINCT user_id) FROM user_activity "
            "WHERE activity_date >= ? GROUP BY activity_date",
            (start.isoformat(),)
        )
        dau_map = dict(cursor.fetchall())

        # Daily signups from users.created_at
        cursor.execute(
            "SELECT DATE(created_at), COUNT(*) FROM users "
            "WHERE DATE(created_at) >= ? GROUP BY DATE(created_at)",
            (start.isoformat(),)
        )
        signup_map = dict(cursor.fetchall())

        # Daily transactions
        cursor.execute(
            "SELECT date, COUNT(*) FROM transactions "
            "WHERE date >= ? GROUP BY date",
            (start.isoformat(),)
        )
        tx_map = dict(cursor.fetchall())

    # Gap-fill
    daily_dau = [{"date": d, "value": dau_map.get(d, 0)} for d in date_range]
    daily_signups = [{"date": d, "value": signup_map.get(d, 0)} for d in date_range]
    daily_transactions = [{"date": d, "value": tx_map.get(d, 0)} for d in date_range]

    # Week-over-week comparisons
    def wow(series):
        this_week = sum(p["value"] for p in series[-7:])
        last_week = sum(p["value"] for p in series[-14:-7])
        if last_week == 0:
            return 100.0 if this_week > 0 else 0.0
        return round((this_week - last_week) / last_week * 100, 1)

    return {
        "daily_active_users": daily_dau,
        "daily_signups": daily_signups,
        "daily_transactions": daily_transactions,
        "comparisons": {
            "dau": wow(daily_dau),
            "signups": wow(daily_signups),
            "transactions": wow(daily_transactions),
        },
    }


@app.get("/api/admin/health")
async def get_admin_health(
    admin_id: int = Depends(get_admin_user_id),
    db: EncryptedDatabase = Depends(get_db)
):
    """
    System health metrics for admin dashboard.
    Protected by admin auth.
    """
    import sqlite3 as _sqlite3
    import shutil

    db_file = Path(db.db_path)
    db_size = db_file.stat().st_size if db_file.exists() else 0

    row_counts = {}
    with _sqlite3.connect(db.db_path) as conn:
        cursor = conn.cursor()
        for table in ["users", "settings", "expenses", "transactions", "user_activity"]:
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                row_counts[table] = cursor.fetchone()[0]
            except Exception:
                row_counts[table] = 0

    # Disk usage - check /data (Fly.io volume) or fall back to DB parent dir
    disk_path = "/data" if Path("/data").exists() else str(db_file.parent)
    disk = shutil.disk_usage(disk_path)

    return {
        "database": {
            "path": str(db_file),
            "size_bytes": db_size,
            "size_mb": round(db_size / (1024 * 1024), 2),
            "row_counts": row_counts,
        },
        "disk": {
            "path": disk_path,
            "total_bytes": disk.total,
            "used_bytes": disk.used,
            "free_bytes": disk.free,
            "used_pct": round(disk.used / disk.total * 100, 1) if disk.total else 0,
        },
    }


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info"
    )
