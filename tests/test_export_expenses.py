"""
Unit tests for expense export functionality.
"""

import pytest
import os
import csv
from pathlib import Path
from src.export_expenses import (
    export_to_csv,
    export_to_excel,
    export_budget_summary
)


class TestCSVExport:
    """Test CSV export functionality."""

    def test_export_to_csv_basic(self, tmp_path):
        """Test basic CSV export."""
        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
            {'name': 'Groceries', 'amount': 300.0, 'is_fixed': False},
        ]

        output_path = tmp_path / "test_export.csv"
        result_path = export_to_csv(expenses, str(output_path))

        assert os.path.exists(result_path)

        # Read and verify content
        with open(result_path, 'r') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        assert len(rows) == 2
        assert rows[0]['name'] == 'Rent'
        assert rows[0]['amount'] == '1500.0'
        assert rows[0]['is_fixed'] == 'yes'
        assert rows[1]['name'] == 'Groceries'
        assert rows[1]['amount'] == '300.0'
        assert rows[1]['is_fixed'] == 'no'

    def test_export_to_csv_auto_filename(self, tmp_path):
        """Test CSV export with auto-generated filename."""
        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
        ]

        # Change to tmp directory for auto-generated file
        original_dir = os.getcwd()
        os.chdir(tmp_path)

        try:
            result_path = export_to_csv(expenses)
            assert os.path.exists(result_path)
            assert result_path.startswith('budget_export_')
            assert result_path.endswith('.csv')
        finally:
            os.chdir(original_dir)

    def test_export_empty_expenses(self, tmp_path):
        """Test exporting empty expense list."""
        expenses = []
        output_path = tmp_path / "empty.csv"
        result_path = export_to_csv(expenses, str(output_path))

        assert os.path.exists(result_path)

        # Should have header but no data rows
        with open(result_path, 'r') as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert len(rows) == 1  # Just header
        assert rows[0] == ['name', 'amount', 'is_fixed']

    def test_export_csv_write_permission_denied(self, tmp_path):
        """Test CSV export when write permission is denied.

        Verifies graceful error handling when user tries to export to
        a location without write permissions.
        """
        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
        ]

        # Try to write to a read-only directory (or invalid path on Windows)
        # Windows: Use a path that definitely doesn't exist and can't be created
        invalid_path = "Z:\\nonexistent\\impossible\\path\\export.csv"

        # Should raise an error (IOError, OSError, or PermissionError)
        with pytest.raises((IOError, OSError, PermissionError, ValueError)):
            export_to_csv(expenses, invalid_path)

    def test_export_csv_invalid_path(self, tmp_path):
        """Test CSV export with invalid file path.

        Tests handling of paths with invalid characters or formats.
        """
        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
        ]

        # Use path with invalid characters for Windows
        invalid_paths = [
            "",  # Empty path
            "   ",  # Whitespace only
        ]

        for invalid_path in invalid_paths:
            with pytest.raises((ValueError, OSError)):
                export_to_csv(expenses, invalid_path)

    def test_export_csv_with_no_data(self, tmp_path):
        """Test exporting when expense data is None.

        Ensures the function handles None gracefully rather than crashing.
        """
        output_path = tmp_path / "none_data.csv"

        # Should raise TypeError or ValueError for None input
        with pytest.raises((TypeError, ValueError)):
            export_to_csv(None, str(output_path))

    def test_export_csv_with_corrupted_expense_data(self, tmp_path):
        """Test export with malformed expense dictionaries.

        Ensures robust handling of expenses missing required fields.
        """
        # Expenses with missing required keys
        malformed_expenses = [
            {'name': 'Rent'},  # Missing 'amount' and 'is_fixed'
            {'amount': 100.0},  # Missing 'name'
            {},  # Empty dict
        ]

        output_path = tmp_path / "corrupted.csv"

        # Should either:
        # 1. Raise an error (preferred)
        # 2. Export with placeholder values
        # We test that it doesn't crash
        try:
            result_path = export_to_csv(malformed_expenses, str(output_path))
            # If it succeeds, verify file was created
            assert os.path.exists(result_path)
        except (KeyError, ValueError, TypeError):
            # If it raises an error, that's also acceptable
            pass


class TestExcelExport:
    """Test Excel export functionality."""

    def test_export_to_excel_basic(self, tmp_path):
        """Test basic Excel export."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
            {'name': 'Groceries', 'amount': 300.0, 'is_fixed': False},
        ]

        output_path = tmp_path / "test_export.xlsx"
        result_path = export_to_excel(expenses, str(output_path))

        assert os.path.exists(result_path)

        # Read and verify content
        wb = openpyxl.load_workbook(result_path)
        ws = wb.active

        # Check header
        assert ws['A1'].value == 'Name'
        assert ws['B1'].value == 'Amount'
        assert ws['C1'].value == 'Type'

        # Check data
        assert ws['A2'].value == 'Rent'
        assert ws['B2'].value == 1500.0
        assert ws['C2'].value == 'Fixed'

        assert ws['A3'].value == 'Groceries'
        assert ws['B3'].value == 300.0
        assert ws['C3'].value == 'Variable'

        # Check total row
        assert ws['A4'].value == 'TOTAL'
        assert ws['B4'].value == 1800.0

        wb.close()

    def test_export_to_excel_auto_filename(self, tmp_path):
        """Test Excel export with auto-generated filename."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
        ]

        original_dir = os.getcwd()
        os.chdir(tmp_path)

        try:
            result_path = export_to_excel(expenses)
            assert os.path.exists(result_path)
            assert result_path.startswith('budget_export_')
            assert result_path.endswith('.xlsx')
        finally:
            os.chdir(original_dir)


class TestBudgetSummaryExport:
    """Test full budget summary export."""

    def test_export_summary_csv_paycheck_mode(self, tmp_path):
        """Test CSV summary export for paycheck mode."""
        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
            {'name': 'Groceries', 'amount': 300.0, 'is_fixed': False},
        ]

        settings = {
            'budget_mode': 'paycheck',
            'monthly_income': 3000.0,
            'days_until_paycheck': 15
        }

        output_path = tmp_path / "summary.csv"
        result_path = export_budget_summary(
            expenses,
            settings,
            str(output_path),
            format='csv'
        )

        assert os.path.exists(result_path)

        # Read content
        with open(result_path, 'r') as f:
            content = f.read()

        # Verify key information is present
        assert 'BUDGET SETTINGS' in content
        assert 'paycheck' in content
        assert '3000' in content
        assert 'MONTHLY EXPENSES' in content
        assert 'Rent' in content
        assert 'Groceries' in content
        assert 'BUDGET SUMMARY' in content

    def test_export_summary_csv_fixed_pool_mode(self, tmp_path):
        """Test CSV summary export for fixed pool mode."""
        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
        ]

        settings = {
            'budget_mode': 'fixed_pool',
            'total_money': 5000.0
        }

        output_path = tmp_path / "summary.csv"
        result_path = export_budget_summary(
            expenses,
            settings,
            str(output_path),
            format='csv'
        )

        assert os.path.exists(result_path)

        with open(result_path, 'r') as f:
            content = f.read()

        assert 'fixed_pool' in content
        assert '5000' in content

    def test_export_summary_excel(self, tmp_path):
        """Test Excel summary export."""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        expenses = [
            {'name': 'Rent', 'amount': 1500.0, 'is_fixed': True},
        ]

        settings = {
            'budget_mode': 'paycheck',
            'monthly_income': 3000.0,
            'days_until_paycheck': 15
        }

        output_path = tmp_path / "summary.xlsx"
        result_path = export_budget_summary(
            expenses,
            settings,
            str(output_path),
            format='excel'
        )

        assert os.path.exists(result_path)

        # Verify it's a valid Excel file
        wb = openpyxl.load_workbook(result_path)
        ws = wb.active

        # Check some key cells
        assert 'BUDGET SETTINGS' in str(ws['A1'].value)
        assert 'MONTHLY EXPENSES' in str(ws['A5'].value) or 'MONTHLY EXPENSES' in str(ws['A6'].value)

        wb.close()
