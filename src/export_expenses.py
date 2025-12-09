"""
Expense export functionality for The Number app.

Allows users to export their budget data to CSV or Excel files.
"""

import csv
import os
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path


def export_to_csv(expenses: List[Dict], output_path: str = None) -> str:
    """
    Export expenses to a CSV file.

    Args:
        expenses: List of expense dictionaries from database
        output_path: Path to save CSV file (auto-generated if None)

    Returns:
        Path to created CSV file
    """
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"budget_export_{timestamp}.csv"

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Write header
        writer.writerow(['name', 'amount', 'is_fixed'])

        # Write expenses
        for exp in expenses:
            is_fixed_str = 'yes' if exp['is_fixed'] else 'no'
            writer.writerow([exp['name'], exp['amount'], is_fixed_str])

    return output_path


def export_to_excel(expenses: List[Dict], output_path: str = None) -> str:
    """
    Export expenses to an Excel file.

    Args:
        expenses: List of expense dictionaries from database
        output_path: Path to save Excel file (auto-generated if None)

    Returns:
        Path to created Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl library required. Run: pip install openpyxl")

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"budget_export_{timestamp}.xlsx"

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Expenses"

    # Style the header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header
    headers = ['Name', 'Amount', 'Type']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    total_amount = 0
    for row, exp in enumerate(expenses, start=2):
        ws.cell(row=row, column=1, value=exp['name'])
        ws.cell(row=row, column=2, value=exp['amount'])
        ws.cell(row=row, column=3, value='Fixed' if exp['is_fixed'] else 'Variable')
        total_amount += exp['amount']

    # Add total row
    if expenses:
        total_row = len(expenses) + 2
        ws.cell(row=total_row, column=1, value='TOTAL')
        ws.cell(row=total_row, column=2, value=total_amount)

        # Style total row
        for col in range(1, 4):
            cell = ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_path)
    return output_path


def export_budget_summary(
    expenses: List[Dict],
    settings: Dict,
    output_path: str = None,
    format: str = 'csv'
) -> str:
    """
    Export complete budget summary including settings and calculated values.

    Args:
        expenses: List of expense dictionaries
        settings: Dictionary of budget settings
        output_path: Path to save file
        format: 'csv' or 'excel'

    Returns:
        Path to created file
    """
    if format.lower() == 'excel':
        return _export_summary_excel(expenses, settings, output_path)
    else:
        return _export_summary_csv(expenses, settings, output_path)


def _export_summary_csv(
    expenses: List[Dict],
    settings: Dict,
    output_path: str = None
) -> str:
    """Export budget summary to CSV."""
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"budget_summary_{timestamp}.csv"

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Budget settings
        writer.writerow(['BUDGET SETTINGS'])
        writer.writerow(['Budget Mode', settings.get('budget_mode', 'N/A')])

        if settings.get('budget_mode') == 'paycheck':
            writer.writerow(['Monthly Income', f"${settings.get('monthly_income', 0):.2f}"])
            writer.writerow(['Days Until Paycheck', settings.get('days_until_paycheck', 0)])
        elif settings.get('budget_mode') == 'fixed_pool':
            writer.writerow(['Total Money', f"${settings.get('total_money', 0):.2f}"])

        writer.writerow([])

        # Expenses
        writer.writerow(['MONTHLY EXPENSES'])
        writer.writerow(['Name', 'Amount', 'Type'])

        total = 0
        for exp in expenses:
            exp_type = 'Fixed' if exp['is_fixed'] else 'Variable'
            writer.writerow([exp['name'], f"${exp['amount']:.2f}", exp_type])
            total += exp['amount']

        writer.writerow(['TOTAL', f"${total:.2f}", ''])
        writer.writerow([])

        # Summary
        writer.writerow(['BUDGET SUMMARY'])
        writer.writerow(['Total Monthly Expenses', f"${total:.2f}"])

        if settings.get('budget_mode') == 'paycheck':
            income = settings.get('monthly_income', 0)
            days = settings.get('days_until_paycheck', 1)
            remaining = income - total
            daily = remaining / days if days > 0 else 0

            writer.writerow(['Monthly Income', f"${income:.2f}"])
            writer.writerow(['Remaining After Expenses', f"${remaining:.2f}"])
            writer.writerow(['Daily Budget', f"${daily:.2f}"])

        writer.writerow([])
        writer.writerow(['Exported', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    return output_path


def _export_summary_excel(
    expenses: List[Dict],
    settings: Dict,
    output_path: str = None
) -> str:
    """Export budget summary to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl library required. Run: pip install openpyxl")

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"budget_summary_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Summary"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    row = 1

    # Budget Settings Section
    ws.cell(row=row, column=1, value="BUDGET SETTINGS")
    ws.cell(row=row, column=1).font = section_font
    row += 1

    ws.cell(row=row, column=1, value="Budget Mode")
    ws.cell(row=row, column=2, value=settings.get('budget_mode', 'N/A').title())
    row += 1

    if settings.get('budget_mode') == 'paycheck':
        ws.cell(row=row, column=1, value="Monthly Income")
        ws.cell(row=row, column=2, value=settings.get('monthly_income', 0))
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1

        ws.cell(row=row, column=1, value="Days Until Paycheck")
        ws.cell(row=row, column=2, value=settings.get('days_until_paycheck', 0))
        row += 1
    elif settings.get('budget_mode') == 'fixed_pool':
        ws.cell(row=row, column=1, value="Total Money")
        ws.cell(row=row, column=2, value=settings.get('total_money', 0))
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1

    row += 1

    # Expenses Section
    ws.cell(row=row, column=1, value="MONTHLY EXPENSES")
    ws.cell(row=row, column=1).font = section_font
    row += 1

    # Header row for expenses
    for col, header in enumerate(['Name', 'Amount', 'Type'], start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Expenses data
    total = 0
    for exp in expenses:
        ws.cell(row=row, column=1, value=exp['name'])
        ws.cell(row=row, column=2, value=exp['amount'])
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=3, value='Fixed' if exp['is_fixed'] else 'Variable')
        total += exp['amount']
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=total)
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    ws.cell(row=row, column=2).font = Font(bold=True)
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = total_fill
    row += 2

    # Summary Section
    ws.cell(row=row, column=1, value="BUDGET SUMMARY")
    ws.cell(row=row, column=1).font = section_font
    row += 1

    ws.cell(row=row, column=1, value="Total Monthly Expenses")
    ws.cell(row=row, column=2, value=total)
    ws.cell(row=row, column=2).number_format = '$#,##0.00'
    row += 1

    if settings.get('budget_mode') == 'paycheck':
        income = settings.get('monthly_income', 0)
        days = settings.get('days_until_paycheck', 1)
        remaining = income - total
        daily = remaining / days if days > 0 else 0

        ws.cell(row=row, column=1, value="Monthly Income")
        ws.cell(row=row, column=2, value=income)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1

        ws.cell(row=row, column=1, value="Remaining After Expenses")
        ws.cell(row=row, column=2, value=remaining)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        row += 1

        ws.cell(row=row, column=1, value="Daily Budget")
        ws.cell(row=row, column=2, value=daily)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Exported")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    return output_path


def get_export_format_choice() -> str:
    """
    Ask user for export format preference.

    Returns:
        'csv' or 'excel'
    """
    print("\n  Export Format:")
    print("    1. CSV (compatible with all spreadsheet apps)")
    print("    2. Excel (formatted with colors and totals)\n")

    while True:
        choice = input("  Select format (1 or 2): ").strip()
        if choice == '1':
            return 'csv'
        elif choice == '2':
            return 'excel'
        else:
            print("  Invalid choice. Please enter 1 or 2.")
